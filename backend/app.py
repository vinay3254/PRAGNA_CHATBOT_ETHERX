"""
Clean Multilingual Chatbot Backend
Simple Flask API with Groq integration
"""
import logging
import requests
import urllib.parse
import json
import os
import io
import time
from datetime import datetime, timezone
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from llm_service import LLMService
import config
from services import memory_db
from services.cache_service import get_cache_service
from services.rag_service import get_rag_service, initialize_rag_with_defaults
from services.web_scraper import update_rag_with_custom_content, update_rag_with_topic_pack
from services.rag_scheduler import start_scheduler, get_rag_update_scheduler, get_scheduler_status, force_rag_update
from services.orchestrator import AIOrchestrator
from services.realtime_events import get_live_feed, get_geo_summary
from services.llm import list_available_models
from services import vision_service
from auth import auth_service, require_auth
from database import db
from chat_management_api import chat_management_bp

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# DEBUG: Log config values immediately after import
logger.info(f"🔥 CONFIG LOADED:")
logger.info(f"   LLM_PROVIDER = '{config.LLM_PROVIDER}'")
logger.info(f"   OLLAMA_ENABLED = {config.OLLAMA_ENABLED}")
logger.info(f"   DEFAULT_MODEL_KEY = {config.DEFAULT_MODEL_KEY}")
logger.info(f"   DEVELOPMENT_MODE = {config.DEVELOPMENT_MODE}")

WORLD_MONITOR_URL = "https://www.worldmonitor.app/"
MAX_UPLOAD_FILES = 100
MAX_UPLOAD_FILE_BYTES = 5 * 1024 * 1024
MAX_UPLOAD_TOTAL_BYTES = 30 * 1024 * 1024
MAX_TEXT_SNIPPET_CHARS = 4000
MAX_COMBINED_TEXT_CHARS = 18000

TEXT_FILE_EXTENSIONS = {
    '.txt', '.md', '.markdown', '.csv', '.tsv', '.json', '.jsonl', '.xml', '.html',
    '.css', '.js', '.jsx', '.ts', '.tsx', '.py', '.java', '.go', '.rs', '.c', '.cpp',
    '.h', '.hpp', '.sql', '.yml', '.yaml', '.toml', '.ini', '.cfg', '.conf', '.log'
}

RICH_TEXT_EXTENSIONS = {'.pdf', '.docx', '.xlsx'}


def _format_bytes(size):
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def _can_extract_text(filename, content_type):
    ext = os.path.splitext(filename.lower())[1]
    if ext in TEXT_FILE_EXTENSIONS:
        return True
    if ext in RICH_TEXT_EXTENSIONS:
        return True
    return bool(content_type and content_type.startswith('text/'))


def _decode_text_bytes(blob):
    for encoding in ('utf-8', 'utf-16', 'latin-1'):
        try:
            return blob.decode(encoding)
        except Exception:
            continue
    return ''


def _extract_text_from_blob(filename, content_type, blob):
    ext = os.path.splitext(filename.lower())[1]

    if ext in TEXT_FILE_EXTENSIONS or (content_type and content_type.startswith('text/')):
        text = _decode_text_bytes(blob)
        return text, 'plain_text', None

    if ext == '.pdf':
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(blob))
            pages = []
            for page in reader.pages[:20]:
                pages.append(page.extract_text() or '')
            return '\n'.join(pages).strip(), 'pdf', None
        except Exception as exc:
            return '', 'pdf', f"PDF extraction unavailable: {exc}"

    if ext == '.docx':
        try:
            from docx import Document

            doc = Document(io.BytesIO(blob))
            text = '\n'.join(p.text for p in doc.paragraphs if p.text)
            return text.strip(), 'docx', None
        except Exception as exc:
            return '', 'docx', f"DOCX extraction unavailable: {exc}"

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
            lines = []
            for sheet in wb.worksheets[:5]:
                lines.append(f"[Sheet: {sheet.title}]")
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_idx > 150:
                        break
                    values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if values:
                        lines.append(' | '.join(values))
            return '\n'.join(lines).strip(), 'xlsx', None
        except Exception as exc:
            return '', 'xlsx', f"XLSX extraction unavailable: {exc}"

    return '', 'metadata_only', None


def _extract_upload_payload(files, relative_paths, attachment_types):
    warnings = []
    file_summaries = []
    text_sections = []
    total_bytes = 0
    skipped_files = 0

    if len(files) > MAX_UPLOAD_FILES:
        warnings.append(
            f"Received {len(files)} files. Only first {MAX_UPLOAD_FILES} will be analyzed."
        )
        files = files[:MAX_UPLOAD_FILES]

    for index, file_storage in enumerate(files):
        filename = (file_storage.filename or '').strip()
        rel_path = relative_paths[index] if index < len(relative_paths) else filename
        attachment_type = attachment_types[index] if index < len(attachment_types) else 'file'

        if not filename:
            skipped_files += 1
            continue

        blob = file_storage.read(MAX_UPLOAD_FILE_BYTES + 1)
        size = len(blob)

        if size > MAX_UPLOAD_FILE_BYTES:
            skipped_files += 1
            warnings.append(
                f"Skipped '{rel_path}' because it exceeds {_format_bytes(MAX_UPLOAD_FILE_BYTES)} per-file limit."
            )
            continue

        if total_bytes + size > MAX_UPLOAD_TOTAL_BYTES:
            skipped_files += 1
            warnings.append(
                f"Skipped '{rel_path}' to stay within total upload limit {_format_bytes(MAX_UPLOAD_TOTAL_BYTES)}."
            )
            continue

        total_bytes += size
        content_type = file_storage.mimetype or 'application/octet-stream'
        summary = {
            'name': filename,
            'relative_path': rel_path,
            'type': attachment_type,
            'content_type': content_type,
            'size_bytes': size,
            'size_human': _format_bytes(size),
            'text_extracted': False,
        }

        # ── Text extraction (documents / plain text) ──────────────────────────
        if _can_extract_text(filename, content_type):
            text_content, extraction_method, extraction_warning = _extract_text_from_blob(
                filename,
                content_type,
                blob,
            )
            summary['extraction_method'] = extraction_method
            if extraction_warning:
                warnings.append(f"{rel_path}: {extraction_warning}")
            if text_content:
                snippet = text_content[:MAX_TEXT_SNIPPET_CHARS]
                summary['text_extracted'] = True
                summary['text_chars'] = len(text_content)
                text_sections.append({
                    'path': rel_path,
                    'content': snippet,
                    'truncated': len(text_content) > len(snippet),
                    'section_type': 'document',
                })

        # ── Vision analysis (PNG / JPG / SVG) ─────────────────────────────────
        elif vision_service.is_supported_image(filename):
            try:
                logger.info("Running vision analysis on '%s' (%s)", filename, content_type)
                vision_description = vision_service.analyze_image(blob, filename)
                summary['text_extracted'] = True
                summary['extraction_method'] = 'vision_analysis'
                summary['text_chars'] = len(vision_description)
                text_sections.append({
                    'path': rel_path,
                    'content': vision_description,
                    'truncated': False,
                    'section_type': 'image_analysis',
                })
                logger.info("Vision analysis complete for '%s'", filename)
            except Exception as vision_err:
                logger.error("Vision analysis failed for '%s': %s", filename, vision_err)
                warnings.append(
                    f"'{rel_path}': image analysis failed — {vision_err}. "
                    "The model will only see filename and file size."
                )

        file_summaries.append(summary)

    return {
        'file_summaries': file_summaries,
        'text_sections': text_sections,
        'warnings': warnings,
        'total_bytes': total_bytes,
        'skipped_files': skipped_files,
        'processed_files': len(file_summaries),
    }


def _build_upload_analysis_prompt(user_message, payload):
    file_lines = []
    for item in payload['file_summaries']:
        extraction_note = 'text extracted' if item.get('text_extracted') else 'metadata only'
        file_lines.append(
            f"- {item['relative_path']} | {item['content_type']} | {item['size_human']} | {extraction_note}"
        )

    sections = []
    used_chars = 0
    for section in payload['text_sections']:
        is_image = section.get('section_type') == 'image_analysis'
        if is_image:
            heading = f"\n[Image Analysis: {section['path']}]\n"
        else:
            heading = f"\n[File: {section['path']}]\n"
        content = section['content']
        chunk = heading + content
        if used_chars + len(chunk) > MAX_COMBINED_TEXT_CHARS:
            break
        used_chars += len(chunk)
        if section.get('truncated'):
            chunk += "\n[Snippet truncated]"
        sections.append(chunk)

    request_text = user_message or 'Analyze these uploaded items and provide the most useful output.'
    warning_text = "\n".join(payload['warnings']) if payload['warnings'] else 'None'
    extracted_text = "\n\n".join(sections) if sections else "No text could be extracted from uploads."

    return (
        "You are analyzing user-uploaded files and folders.\n"
        f"User request: {request_text}\n\n"
        "Uploaded items:\n"
        f"{chr(10).join(file_lines) if file_lines else '- None'}\n\n"
        "Warnings and constraints:\n"
        f"{warning_text}\n\n"
        "Extracted content and image descriptions:\n"
        f"{extracted_text}\n\n"
        "Instructions:\n"
        "1) Directly answer the user's request using the uploaded content and image descriptions above.\n"
        "2) For images, use the [Image Analysis] section which contains a visual description of what the image shows.\n"
        "3) Clearly separate confirmed findings from assumptions.\n"
        "4) If information is missing, say exactly what is missing and suggest next steps.\n"
        "5) Keep the output practical and action-oriented."
    )


def _build_image_generation_prompt(user_prompt: str, style: str, quality: str) -> str:
    style_map = {
        'photo': 'photorealistic, natural lighting, realistic textures, high-detail camera-like output',
        'cinematic': 'cinematic composition, dramatic lighting, rich color grading, movie still quality',
        'illustration': 'clean illustration style, crisp shapes, deliberate color palette, high detail',
        'concept_art': 'concept-art quality, strong silhouette language, production-ready scene design',
        'product': 'product-shot framing, studio lighting, premium commercial look, clean background',
    }
    style_hint = style_map.get(style, style_map['cinematic'])
    detail_hint = 'ultra detailed' if quality == 'hd' else 'high quality'
    return (
        f"{user_prompt.strip()}\n\n"
        f"Render style guidance: {style_hint}.\n"
        f"Quality target: {detail_hint}.\n"
        "Requirements: keep subject accurate to prompt, avoid unwanted text artifacts, maintain coherent anatomy and perspective."
    )


def _extract_first_image_url(payload):
    if isinstance(payload, str):
        if payload.startswith('http://') or payload.startswith('https://') or payload.startswith('data:image/'):
            return payload
        return None

    if isinstance(payload, list):
        for item in payload:
            found = _extract_first_image_url(item)
            if found:
                return found
        return None

    if not isinstance(payload, dict):
        return None

    for key in ('image', 'image_url', 'url'):
        value = payload.get(key)
        found = _extract_first_image_url(value)
        if found:
            return found

    for key in ('output', 'result', 'results', 'data', 'images', 'artifacts'):
        value = payload.get(key)
        found = _extract_first_image_url(value)
        if found:
            return found

    b64 = payload.get('b64_json') if isinstance(payload, dict) else None
    if isinstance(b64, str) and b64.strip():
        return f"data:image/png;base64,{b64.strip()}"

    return None


def _normalize_language_code(language_value):
    raw = str(language_value or 'en').strip().lower().replace('_', '-')
    if not raw:
        return 'en'

    aliases = {
        'english': 'en',
        'hindi': 'hi',
        'tamil': 'ta',
        'telugu': 'te',
        'kannada': 'kn',
        'malayalam': 'ml',
        'marathi': 'mr',
        'gujarati': 'gu',
        'punjabi': 'pa',
        'bengali': 'bn',
        'bangla': 'bn',
        'urdu': 'ur',
    }

    if raw in config.SUPPORTED_LANGUAGES:
        return raw

    if raw in aliases:
        return aliases[raw]

    base = raw.split('-')[0]
    if base in config.SUPPORTED_LANGUAGES:
        return base

    if base in aliases:
        return aliases[base]

    return 'en'


def _generate_with_runway(prompt: str, size: str):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {config.RUNWAY_API_KEY}',
        'X-Runway-Version': '2024-11-06',
    }
    ratio_map = {
        '1024x1024': '1024:1024',
        '1024x1536': '1080:1440',
        '1536x1024': '1440:1080',
    }
    ratio = ratio_map.get(size, '1024:1024')
    payload = {
        'model': config.RUNWAY_MODEL,
        'promptText': prompt,
        'ratio': ratio,
    }

    create_res = requests.post(
        f"{config.RUNWAY_API_URL.rstrip('/')}/text_to_image",
        headers=headers,
        json=payload,
        timeout=config.RUNWAY_TIMEOUT,
    )

    if create_res.status_code >= 400:
        raise RuntimeError(f"Runway request failed ({create_res.status_code}): {create_res.text[:300]}")

    create_data = create_res.json()
    direct_image = _extract_first_image_url(create_data)
    if direct_image:
        return direct_image

    task_id = create_data.get('id') or create_data.get('taskId') or create_data.get('task_id')
    if not task_id:
        raise RuntimeError('Runway response did not include image or task id')

    for _ in range(max(1, config.RUNWAY_POLL_ATTEMPTS)):
        poll_res = requests.get(
            f"{config.RUNWAY_API_URL.rstrip('/')}/tasks/{task_id}",
            headers=headers,
            timeout=config.RUNWAY_TIMEOUT,
        )
        if poll_res.status_code >= 400:
            raise RuntimeError(f"Runway task polling failed ({poll_res.status_code}): {poll_res.text[:300]}")

        poll_data = poll_res.json()
        status = str(poll_data.get('status', '')).lower()

        if status in {'succeeded', 'completed', 'success'}:
            image = _extract_first_image_url(poll_data)
            if image:
                return image
            raise RuntimeError('Runway task completed but no image payload was found')

        if status in {'failed', 'cancelled', 'canceled'}:
            raise RuntimeError(f"Runway task failed: {json.dumps(poll_data)[:400]}")

        time.sleep(max(0.2, config.RUNWAY_POLL_INTERVAL_SECONDS))

    raise RuntimeError('Runway image generation timed out')

# Initialize Flask app
app = Flask(__name__, static_folder='static')
CORS(app)

# Initialize LLM service
llm = LLMService()
orchestrator = AIOrchestrator(llm)

# Initialize RAG service with default knowledge base (skip if it takes too long)
logger.info("🚀 Initializing RAG service...")
rag = get_rag_service()
try:
    import threading
    import queue
    
    rag_initialized = queue.Queue()
    
    def init_rag_with_timeout():
        try:
            if initialize_rag_with_defaults():
                llm.enable_rag()
                rag_initialized.put(True)
            else:
                rag_initialized.put(False)
        except Exception as e:
            logger.warning(f"RAG init error: {e}")
            rag_initialized.put(False)
    
    # Run RAG initialization in background thread with 10 second timeout
    rag_thread = threading.Thread(target=init_rag_with_timeout, daemon=True)
    rag_thread.start()
    rag_thread.join(timeout=10)
    
    try:
        success = rag_initialized.get_nowait()
        if success:
            logger.info("✅ RAG service ready with knowledge base")
        else:
            logger.warning("⚠️ RAG service initialization failed")
    except queue.Empty:
        logger.warning("⚠️ RAG initialization timed out (taking too long); continuing without it")
        
except Exception as e:
    logger.warning(f"⚠️ RAG service not available: {e}")

# Start background scheduler for RAG updates (in background thread)
logger.info("🔄 Starting RAG update scheduler...")
try:
    start_scheduler()
    scheduler_status = get_scheduler_status()
    logger.info(f"📊 Scheduler status: {scheduler_status}")
except Exception as e:
    logger.warning(f"⚠️ RAG scheduler error: {e}")

logger.info("✅ Chatbot server starting...")
logger.info(f"✅ Using Groq model: {config.GROQ_MODEL}")
logger.info(f"✅ Supported languages: {list(config.SUPPORTED_LANGUAGES.keys())}")

# Public endpoints that don't require authentication
PUBLIC_ENDPOINTS = [
    '/api/chat',
    '/api/orchestrator/query',
    '/api/orchestrator/analyze_uploads',
    '/api/models/catalog',
    '/api/platform/status',
    '/api/world-monitor/config',
    '/api/clear_history',
    '/api/status',
    '/api/events/feed',
    '/api/dashboard/geo',
    '/api/cache/stats',
    '/api/cache/cleanup',
    '/api/rag/stats',
    '/api/health',
    '/api/test-ollama',
    '/api/tts',
    '/api/translate',
    '/api/summarize',
    '/api/auth/login',
    '/api/auth/register',
]

@app.before_request
def before_request():
    """Allow public endpoints without auth"""
    if request.path in PUBLIC_ENDPOINTS or request.path.startswith('/api/'):
        # Public endpoints bypass auth
        return None
    return None

@app.route('/')
def index():
    """Serve the frontend"""
    return send_from_directory('static', 'index.html')


@app.route('/api/chat', methods=['POST'])
def chat():
    """
    Main chat endpoint
    Expects JSON: {"message": "user message", "language": "en", "user_id": "optional", "chat_mode": "general"}
    Returns JSON: {"response": "ai response"}
    """
    try:
        data = request.json
        
        # Validate input
        if not data or 'message' not in data:
            return jsonify({'error': 'Message is required'}), 400
        
        user_message = data.get('message', '').strip()
        language = _normalize_language_code(data.get('language', 'en'))
        user_id = data.get('user_id', 'default')
        chat_mode = data.get('chat_mode', 'general')
        model_override = data.get('model_override')
        fallback_models = data.get('fallback_models')
        
        if not user_message:
            return jsonify({'error': 'Message cannot be empty'}), 400
        
        logger.info(f"Received message: {user_message[:50]}... (language: {language}, mode: {chat_mode})")
        
        # Unified orchestration path (agent tools + LLM/RAG)
        result = orchestrator.handle_query(
            user_message,
            language=language,
            user_id=user_id,
            chat_mode=chat_mode,
            model_override=model_override,
            fallback_models=fallback_models,
        )
        
        return jsonify({
            'response': result['response'],
            'language': result['language'],
            'chat_mode': result['chat_mode'],
            'route': result['route'],
            'action': result['action'],
            'actions': result['actions'],
            'web_search_sources': result['web_search_sources']
        })
        
    except Exception as e:
        logger.error(f"Error in chat endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/debug/config', methods=['GET'])
def debug_config():
    """Debug endpoint to check Flask app configuration"""
    return jsonify({
        'LLM_PROVIDER': config.LLM_PROVIDER,
        'OLLAMA_ENABLED': config.OLLAMA_ENABLED,
        'OLLAMA_API_URL': config.OLLAMA_API_URL,
        'OLLAMA_MODEL': config.OLLAMA_MODEL,
        'DEFAULT_MODEL_KEY': config.DEFAULT_MODEL_KEY,
        'DEVELOPMENT_MODE': config.DEVELOPMENT_MODE,
    })


@app.route('/api/test/ollama-direct', methods=['POST'])
def test_ollama_direct():
    """Test Ollama directly, bypassing orchestrator"""
    from services.llm import generate_completion
    
    try:
        data = request.json or {}
        message = data.get('message', 'Hello')
        
        print(f"📍 TEST ENDPOINT: Calling generate_completion directly", flush=True)
        messages = [{"role": "user", "content": message}]
        response = generate_completion(messages)
        print(f"📍 TEST ENDPOINT: Got response: {response[:100]}", flush=True)
        
        return jsonify({
            'success': True,
            'response': response,
            'provider': config.LLM_PROVIDER,
        })
    except Exception as e:
        print(f"📍 TEST ENDPOINT: EXCEPTION: {e}", flush=True)
        return jsonify({
            'success': False,
            'error': str(e),
            'provider': config.LLM_PROVIDER,
        }), 500


@app.route('/api/orchestrator/query', methods=['POST'])
def orchestrator_query():
    """Explicit orchestrator endpoint for agent + RAG + LLM routing."""
    try:
        data = request.json or {}
        user_message = (data.get('message') or '').strip()
        language = _normalize_language_code(data.get('language', 'en'))
        user_id = data.get('user_id', 'default')
        chat_mode = data.get('chat_mode', 'general')
        model_override = data.get('model_override')
        fallback_models = data.get('fallback_models')

        if not user_message:
            return jsonify({'error': 'Message is required'}), 400

        result = orchestrator.handle_query(
            user_message,
            language=language,
            user_id=user_id,
            chat_mode=chat_mode,
            model_override=model_override,
            fallback_models=fallback_models,
        )
        return jsonify(result)

    except Exception as e:
        logger.error(f"Error in orchestrator_query endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/orchestrator/analyze_uploads', methods=['POST'])
def orchestrator_analyze_uploads():
    """Analyze uploaded files/folders/photos and generate an orchestrated response."""
    try:
        files = request.files.getlist('files')
        relative_paths = request.form.getlist('relative_paths')
        attachment_types = request.form.getlist('attachment_types')

        user_message = (request.form.get('message') or '').strip()
        language = _normalize_language_code(request.form.get('language', 'en'))
        user_id = request.form.get('user_id', 'default')
        chat_mode = request.form.get('chat_mode', 'general')
        model_override = request.form.get('model_override')
        fallback_models_raw = request.form.get('fallback_models', '[]')

        try:
            fallback_models = json.loads(fallback_models_raw)
            if not isinstance(fallback_models, list):
                fallback_models = []
        except Exception:
            fallback_models = []

        if not files and not user_message:
            return jsonify({'error': 'Upload at least one file or provide a message'}), 400

        payload = _extract_upload_payload(files, relative_paths, attachment_types)
        if payload['processed_files'] == 0 and not user_message:
            return jsonify({'error': 'No valid files to analyze'}), 400

        composed_prompt = _build_upload_analysis_prompt(user_message, payload)

        result = orchestrator.handle_query(
            composed_prompt,
            language=language,
            user_id=user_id,
            chat_mode=chat_mode,
            model_override=model_override,
            fallback_models=fallback_models,
        )

        result['upload_analysis'] = {
            'processed_files': payload['processed_files'],
            'skipped_files': payload['skipped_files'],
            'total_size_bytes': payload['total_bytes'],
            'warnings': payload['warnings'],
            'files': payload['file_summaries'],
        }

        return jsonify(result)

    except Exception as e:
        logger.error(f"Error in orchestrator_analyze_uploads endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Failed to analyze uploads'}), 500


@app.route('/api/clear_history', methods=['POST'])
def clear_history():
    """Clear conversation history for a user"""
    try:
        data = request.json or {}
        user_id = data.get('user_id', 'default')
        
        llm.clear_history(user_id)
        
        return jsonify({'message': 'History cleared'})
        
    except Exception as e:
        logger.error(f"Error clearing history: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/health', methods=['GET'])
@app.route('/api/health', methods=['GET'])
def health_check():
    """
    Comprehensive health check endpoint to test all critical systems.
    Tests: API key configuration, LLM connectivity, database, cache, RAG
    """
    health_status = {
        'status': 'unknown',
        'timestamp': time.time(),
        'systems': {},
        'errors': []
    }
    
    # Check 1: API Key Configuration
    try:
        groq_key_configured = bool(config.GROQ_API_KEY and len(config.GROQ_API_KEY.strip()) > 0)
        openai_key_configured = bool(config.OPENAI_API_KEY and len(config.OPENAI_API_KEY.strip()) > 0)
        ollama_configured = config.OLLAMA_ENABLED
        
        if not (groq_key_configured or openai_key_configured or ollama_configured):
            health_status['errors'].append(
                "❌ No LLM provider configured. Set one of: GROQ_API_KEY, OPENAI_API_KEY, or OLLAMA_ENABLED=True"
            )
        
        health_status['systems']['api_keys'] = {
            'groq_configured': groq_key_configured,
            'groq_key_preview': config.GROQ_API_KEY[:10] + '...' if groq_key_configured else 'NOT SET',
            'openai_configured': openai_key_configured,
            'ollama_enabled': ollama_configured,
            'development_mode': config.DEVELOPMENT_MODE
        }
    except Exception as e:
        health_status['systems']['api_keys'] = {'error': str(e)}
        health_status['errors'].append(f'❌ API key check failed: {e}')
    
    # Check 2: LLM Test Request
    try:
        from services.llm import _request_completion
        test_messages = [
            {'role': 'system', 'content': 'You are a helpful assistant. Respond with exactly: SUCCESS'},
            {'role': 'user', 'content': 'test'}
        ]
        
        try:
            response = _request_completion(test_messages, config.DEFAULT_MODEL_KEY)
            if response and 'success' in response.lower():
                health_status['systems']['llm'] = {
                    'status': 'healthy',
                    'model': config.DEFAULT_MODEL_KEY,
                    'response_length': len(response)
                }
            else:
                health_status['systems']['llm'] = {
                    'status': 'responding',
                    'model': config.DEFAULT_MODEL_KEY,
                    'response_sample': response[:100]
                }
        except RuntimeError as e:
            if '401' in str(e) or 'not configured' in str(e).lower():
                health_status['errors'].append(f'❌ LLM API Error: {e}')
                health_status['systems']['llm'] = {'status': 'error', 'error': str(e)[:200]}
            else:
                raise
    except Exception as e:
        health_status['systems']['llm'] = {'status': 'error', 'error': str(e)[:200]}
        health_status['errors'].append(f'❌ LLM check failed: {str(e)[:100]}')
    
    # Check 3: Database
    try:
        conn = db.get_connection()
        conn.execute('SELECT 1')
        conn.close()
        health_status['systems']['database'] = {'status': 'healthy'}
    except Exception as e:
        health_status['systems']['database'] = {'status': 'error', 'error': str(e)[:100]}
        health_status['errors'].append(f'❌ Database error: {str(e)[:100]}')
    
    # Check 4: Cache
    try:
        cache = get_cache_service()
        stats = cache.get_stats()
        health_status['systems']['cache'] = {
            'status': 'healthy',
            'entries': stats.get('total_entries', 0),
            'hit_rate': f"{stats.get('hit_rate_percent', 0):.1f}%"
        }
    except Exception as e:
        health_status['systems']['cache'] = {'status': 'error', 'error': str(e)[:100]}
    
    # Check 5: RAG
    try:
        rag = get_rag_service()
        rag_stats = rag.get_stats()
        health_status['systems']['rag'] = {
            'status': 'healthy' if rag_stats.get('enabled') else 'disabled',
            'documents': rag_stats.get('document_count', 0),
            'enabled': rag_stats.get('enabled', False)
        }
    except Exception as e:
        health_status['systems']['rag'] = {'status': 'error', 'error': str(e)[:100]}
    
    # Overall status determination
    critical_errors = [e for e in health_status['errors'] if '❌' in e]
    if critical_errors:
        health_status['status'] = 'unhealthy'
        http_status = 503
    elif health_status['errors']:
        health_status['status'] = 'degraded'
        http_status = 200
    else:
        health_status['status'] = 'healthy'
        http_status = 200
    
    return jsonify(health_status), http_status


@app.route('/api/status', methods=['GET'])
def status():
    """Compatibility endpoint for frontend health check"""
    return jsonify({
        'status': 'healthy',
        'models_loaded': True,
        'model': config.GROQ_MODEL
    })


@app.route('/api/test-ollama', methods=['POST'])
@app.route('/api/test-ollama', methods=['GET'])
def test_ollama():
    """
    Test endpoint to verify Ollama connectivity and functionality.
    Returns: Connection status, model info, and test response
    """
    logger.info("🧪 Testing Ollama connection...")
    
    test_result = {
        'status': 'unknown',
        'ollama_enabled': config.OLLAMA_ENABLED,
        'ollama_url': config.OLLAMA_API_URL,
        'ollama_model': config.OLLAMA_MODEL,
        'llm_provider': config.LLM_PROVIDER,
        'messages': [],
        'errors': []
    }
    
    # Check 1: Is Ollama enabled?
    if not config.OLLAMA_ENABLED:
        test_result['errors'].append(f"❌ Ollama is DISABLED. Set OLLAMA_ENABLED=True in backend/.env")
        test_result['status'] = 'error'
        return jsonify(test_result), 503
    
    test_result['messages'].append(f"✓ Ollama ENABLED")
    
    # Check 2: Can we reach Ollama?
    ollama_url = config.OLLAMA_API_URL.rstrip('/')
    try:
        response = requests.get(
            f"{ollama_url}/api/tags",
            timeout=5
        )
        response.raise_for_status()
        models = response.json().get('models', [])
        test_result['messages'].append(f"✓ Ollama is REACHABLE at {ollama_url}")
        test_result['available_models'] = [m.get('name', m) for m in models]
        
        # Check 3: Is our model available?
        model_names = [m.get('name', '') if isinstance(m, dict) else str(m) for m in models]
        if config.OLLAMA_MODEL in model_names or any(config.OLLAMA_MODEL in str(m) for m in model_names):
            test_result['messages'].append(f"✓ Model '{config.OLLAMA_MODEL}' is AVAILABLE")
        else:
            test_result['errors'].append(
                f"⚠️  Model '{config.OLLAMA_MODEL}' not found.\n"
                f"Available models: {', '.join(model_names) if model_names else 'none'}\n"
                f"Pull model with: ollama pull {config.OLLAMA_MODEL}"
            )
            test_result['status'] = 'warning'
            return jsonify(test_result), 207  # Multi-status
            
    except requests.exceptions.ConnectionError:
        test_result['errors'].append(
            f"❌ Cannot connect to Ollama at {ollama_url}\n"
            f"1. Start Ollama: ollama run {config.OLLAMA_MODEL}\n"
            f"2. Check OLLAMA_API_URL in backend/.env"
        )
        test_result['status'] = 'error'
        return jsonify(test_result), 503
    except requests.exceptions.Timeout:
        test_result['errors'].append(f"❌ Ollama timeout at {ollama_url}")
        test_result['status'] = 'error'
        return jsonify(test_result), 503
    except Exception as e:
        test_result['errors'].append(f"❌ Error reaching Ollama: {str(e)}")
        test_result['status'] = 'error'
        return jsonify(test_result), 500
    
    # Check 4: Test a simple request
    test_messages = [
        {'role': 'system', 'content': 'You are a helpful assistant. Be concise.'},
        {'role': 'user', 'content': 'Say "Ollama is working!" and nothing else.'}
    ]
    
    try:
        from services.llm import _call_ollama_direct
        test_response = _call_ollama_direct(test_messages)
        
        test_result['messages'].append(f"✓ Test request SUCCESSFUL")
        test_result['test_response'] = test_response[:200]  # First 200 chars
        test_result['status'] = 'healthy'
        
        return jsonify(test_result), 200
        
    except Exception as e:
        test_result['errors'].append(f"❌ Test request failed: {str(e)}")
        test_result['status'] = 'error'
        return jsonify(test_result), 500


@app.route('/api/platform/status', methods=['GET'])
def platform_status():
    """Unified product health summary for dashboard/header indicators."""
    try:
        rag_stats_data = get_rag_service().get_stats()
        scheduler = get_scheduler_status()
        cache_stats_data = get_cache_service().get_stats()

        return jsonify({
            'status': 'success',
            'platform': {
                'model': config.GROQ_MODEL,
                'default_model_key': config.DEFAULT_MODEL_KEY,
                'providers': {
                    'groq_enabled': bool(config.GROQ_API_KEY),
                    'ollama_enabled': config.OLLAMA_ENABLED,
                },
                'rag_enabled': rag_stats_data.get('enabled', False),
                'rag_documents': rag_stats_data.get('document_count', 0),
                'scheduler_running': scheduler.get('running', False),
                'cache_hit_rate_percent': cache_stats_data.get('hit_rate_percent', 0.0),
                'cache_entries': cache_stats_data.get('total_entries', 0),
            }
        })
    except Exception as e:
        logger.error(f"Error in platform_status endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve platform status'}), 500


@app.route('/api/models/catalog', methods=['GET'])
def models_catalog():
    """Expose available open model options and recommendation profiles."""
    return jsonify({
        'status': 'success',
        'default_model_key': config.DEFAULT_MODEL_KEY,
        'fallback_models': config.DEFAULT_MODEL_FALLBACKS,
        'recommendations': config.MODEL_RECOMMENDATIONS,
        'models': list_available_models(),
    })


@app.route('/api/world-monitor/config', methods=['GET'])
def world_monitor_config():
    """Expose World Monitor integration metadata to the frontend dashboard."""
    return jsonify({
        'status': 'success',
        'world_monitor': {
            'url': WORLD_MONITOR_URL,
            'embeddable': False,
            'reason': 'worldmonitor.app sets SAMEORIGIN/frame-ancestors restrictions',
            'integration_mode': 'external_tab',
            'focus': 'india-first',
            'primary_region': 'India',
        }
    })


@app.route('/api/images/generate', methods=['POST'])
def generate_image():
    """Generate an AI image using configured provider (Runway/OpenAI/fallback)."""
    try:
        data = request.json or {}
        prompt = (data.get('prompt') or '').strip()
        style = (data.get('style') or 'cinematic').strip().lower()
        quality = (data.get('quality') or 'hd').strip().lower()
        size = (data.get('size') or '1024x1024').strip().lower()
        requested_provider = (data.get('provider') or config.IMAGE_PROVIDER or 'auto').strip().lower()

        if not prompt:
            return jsonify({'error': 'Prompt is required'}), 400

        if not config.IMAGE_GENERATION_ENABLED:
            return jsonify({'error': 'Image generation is disabled'}), 503

        allowed_sizes = {'1024x1024', '1024x1536', '1536x1024'}
        if size not in allowed_sizes:
            size = '1024x1024'

        if quality not in {'standard', 'hd'}:
            quality = 'hd'

        enhanced_prompt = _build_image_generation_prompt(prompt, style, quality)

        # 1) Runway path (explicit or auto)
        if requested_provider in {'runway', 'auto'} and config.RUNWAY_API_KEY:
            try:
                runway_image = _generate_with_runway(enhanced_prompt, size)
                return jsonify({
                    'status': 'success',
                    'provider': 'runway',
                    'image': runway_image,
                    'model': config.RUNWAY_MODEL,
                    'style': style,
                    'quality': quality,
                    'size': size,
                    'effective_prompt': enhanced_prompt,
                })
            except Exception as runway_exc:
                logger.error(f"Runway image generation failed: {runway_exc}")
                if requested_provider == 'runway' and config.RUNWAY_STRICT_MODE:
                    return jsonify({
                        'error': 'Runway image generation failed',
                        'provider': 'runway',
                        'provider_error': str(runway_exc),
                    }), 502
                logger.warning("Falling back to other image providers after Runway failure")

        # 2) OpenAI path (explicit or auto)
        if not config.OPENAI_API_KEY:
            if not config.IMAGE_FALLBACK_PROVIDER_ENABLED:
                return jsonify({'error': 'Image generation is not configured (missing provider API key)'}), 503

            encoded_prompt = urllib.parse.quote_plus(enhanced_prompt)
            fallback_url = (
                f"{config.IMAGE_FALLBACK_PROVIDER_URL.rstrip('/')}/{encoded_prompt}"
                f"?width={size.split('x')[0]}&height={size.split('x')[1]}&model=flux&nologo=true"
            )
            return jsonify({
                'status': 'success',
                'provider': 'pollinations-fallback',
                'image': fallback_url,
                'model': 'flux',
                'style': style,
                'quality': quality,
                'size': size,
                'effective_prompt': enhanced_prompt,
                'note': 'Fallback provider used because primary provider key is not configured.',
            })

        payload = {
            'model': config.OPENAI_IMAGE_MODEL,
            'prompt': enhanced_prompt,
            'size': size,
            'quality': quality,
            'n': 1,
            'response_format': 'b64_json',
        }

        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {config.OPENAI_API_KEY}',
        }

        response = requests.post(
            'https://api.openai.com/v1/images/generations',
            headers=headers,
            json=payload,
            timeout=config.OPENAI_TIMEOUT,
        )

        if response.status_code >= 400:
            provider_error = response.text[:500]
            logger.error(f"Image generation failed: {response.status_code} {provider_error}")
            return jsonify({
                'error': 'Image generation failed on provider',
                'provider_status': response.status_code,
                'provider_error': provider_error,
            }), 502

        result = response.json()
        data_items = result.get('data') or []
        if not data_items:
            return jsonify({'error': 'No image returned from provider'}), 502

        first = data_items[0]
        image_b64 = first.get('b64_json')
        image_url = first.get('url')

        final_image = image_url or (f"data:image/png;base64,{image_b64}" if image_b64 else None)
        if not final_image:
            return jsonify({'error': 'Provider response missing image payload'}), 502

        return jsonify({
            'status': 'success',
            'provider': 'openai',
            'image': final_image,
            'model': config.OPENAI_IMAGE_MODEL,
            'style': style,
            'quality': quality,
            'size': size,
            'effective_prompt': enhanced_prompt,
        })

    except Exception as e:
        logger.error(f"Error generating image: {e}", exc_info=True)
        return jsonify({'error': 'Failed to generate image'}), 500


@app.route('/api/events/feed', methods=['GET'])
def events_feed():
    """Return normalized realtime event feed for dashboard/chat intelligence panels."""
    try:
        limit = int(request.args.get('limit', 20))
        limit = max(1, min(limit, 100))
        focus = (request.args.get('focus', 'india') or 'india').strip().lower()
        payload = get_live_feed(limit=limit, focus=focus)
        if focus == 'india' and not (payload.get('events') or []):
            payload['events'] = [{
                'event_id': 'india-focus-api-fallback',
                'title': 'India Focus Active',
                'summary': 'India-first mode is enabled. Live providers returned limited data in this cycle; refreshing shortly.',
                'source': 'Pragna Monitor',
                'link': '',
                'published_at': datetime.now(timezone.utc).isoformat(),
                'region': 'India',
                'severity': 'low',
                'coordinates': {'lat': 20.5937, 'lon': 78.9629},
                'is_india_focus': True,
                'india_relevance': 999,
            }]
            payload['count'] = 1
        elif focus == 'india':
            events = payload.get('events') or []
            if events and events[0].get('region') != 'India':
                events.insert(0, {
                    'event_id': 'india-focus-priority-api',
                    'title': 'India Focus Priority',
                    'summary': 'This dashboard is currently prioritized for India. Global headlines are shown after India-priority context.',
                    'source': 'Pragna Monitor',
                    'link': '',
                    'published_at': datetime.now(timezone.utc).isoformat(),
                    'region': 'India',
                    'severity': 'low',
                    'coordinates': {'lat': 20.5937, 'lon': 78.9629},
                    'is_india_focus': True,
                    'india_relevance': 999,
                })
                payload['events'] = events[:limit]
                payload['count'] = len(payload['events'])
        return jsonify({'status': 'success', 'focus': focus, **payload})
    except Exception as e:
        logger.error(f"Error in events_feed endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve events feed'}), 500


@app.route('/api/dashboard/geo', methods=['GET'])
def dashboard_geo():
    """Return region-level event aggregates for map/dashboard rendering."""
    try:
        limit = int(request.args.get('limit', 50))
        limit = max(1, min(limit, 200))
        focus = (request.args.get('focus', 'india') or 'india').strip().lower()
        payload = get_geo_summary(limit=limit, focus=focus)
        if focus == 'india' and not (payload.get('regions') or []):
            payload['regions'] = [{'region': 'India', 'events': 1, 'lat': 20.5937, 'lon': 78.9629}]
        elif focus == 'india':
            regions = payload.get('regions') or []
            if regions and regions[0].get('region') != 'India':
                regions.insert(0, {'region': 'India', 'events': max(1, int(regions[0].get('events', 1))), 'lat': 20.5937, 'lon': 78.9629})
                payload['regions'] = regions[:limit]
        return jsonify({'status': 'success', 'focus': focus, **payload})
    except Exception as e:
        logger.error(f"Error in dashboard_geo endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve dashboard geo summary'}), 500


@app.route('/api/cache/stats', methods=['GET'])
def cache_stats():
    """
    Get cache performance statistics
    Returns: JSON with cache hit rate, operation counts, and memory snapshot
    """
    try:
        cache = get_cache_service()
        stats = cache.get_stats()
        
        logger.info(f"📊 Cache stats requested - Hit rate: {stats['hit_rate_percent']:.1f}%")
        
        return jsonify({
            'status': 'success',
            'cache_hits': stats['total_hits'],
            'cache_misses': stats['total_misses'],
            'cache_sets': stats['total_sets'],
            'hit_rate_percent': stats['hit_rate_percent'],
            'total_entries': stats['total_entries'],
            'approx_memory_mb': stats['approx_memory_mb']
        })
        
    except Exception as e:
        logger.error(f"Error retrieving cache stats: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve cache statistics'}), 500


@app.route('/api/cache/cleanup', methods=['POST'])
def cache_cleanup():
    """
    Manually trigger cache cleanup (remove expired entries)
    Returns: JSON with number of entries cleaned up
    """
    try:
        cache = get_cache_service()
        cleaned_count = cache.cleanup_expired()
        
        logger.info(f"🧹 Cache cleanup performed - Removed {cleaned_count} expired entries")
        
        return jsonify({
            'status': 'success',
            'cleaned_entries': cleaned_count,
            'message': f'Removed {cleaned_count} expired cache entries'
        })
        
    except Exception as e:
        logger.error(f"Error during cache cleanup: {e}", exc_info=True)
        return jsonify({'error': 'Failed to cleanup cache'}), 500


@app.route('/api/rag/stats', methods=['GET'])
def rag_stats():
    """
    Get RAG (Retrieval-Augmented Generation) statistics
    Returns: JSON with RAG status, document count, and model info
    """
    try:
        rag = get_rag_service()
        stats = rag.get_stats()
        
        logger.info(f"📊 RAG stats requested - Enabled: {stats['enabled']}, Documents: {stats['document_count']}")
        
        return jsonify({
            'status': 'success',
            'rag_enabled': stats['enabled'],
            'has_index': stats['has_index'],
            'document_count': stats['document_count'],
            'model': stats['model'],
            'rag_active': llm.use_rag
        })
        
    except Exception as e:
        logger.error(f"Error retrieving RAG stats: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve RAG statistics'}), 500


@app.route('/api/rag/add_documents', methods=['POST'])
def rag_add_documents():
    """
    Add documents to RAG knowledge base
    Expects JSON: {"documents": ["doc1", "doc2", ...], "document_ids": ["id1", "id2", ...] (optional)}
    Returns: JSON with success status
    """
    try:
        data = request.json or {}
        documents = data.get('documents', [])
        document_ids = data.get('document_ids')
        
        if not documents:
            return jsonify({'error': 'No documents provided'}), 400
        
        if not isinstance(documents, list):
            documents = [documents]
        
        rag = get_rag_service()
        if not rag.enabled:
            return jsonify({'error': 'RAG service not available'}), 503
        
        success = rag.add_documents(documents, document_ids)
        
        if success:
            logger.info(f"📚 Added {len(documents)} documents to RAG knowledge base")
            return jsonify({
                'status': 'success',
                'documents_added': len(documents),
                'total_documents': len(rag.metadata)
            })
        else:
            return jsonify({'error': 'Failed to add documents'}), 500
        
    except Exception as e:
        logger.error(f"Error adding RAG documents: {e}", exc_info=True)
        return jsonify({'error': 'Failed to add documents'}), 500


@app.route('/api/rag/clear', methods=['POST'])
def rag_clear():
    """
    Clear all documents from RAG knowledge base
    Returns: JSON with success status
    """
    try:
        rag = get_rag_service()
        success = rag.clear_index()
        
        if success:
            logger.info("🧹 RAG knowledge base cleared")
            return jsonify({
                'status': 'success',
                'message': 'RAG knowledge base cleared'
            })
        else:
            return jsonify({'error': 'Failed to clear RAG knowledge base'}), 500
        
    except Exception as e:
        logger.error(f"Error clearing RAG knowledge base: {e}", exc_info=True)
        return jsonify({'error': 'Failed to clear RAG knowledge base'}), 500


@app.route('/api/rag/update_web_content', methods=['POST'])
def rag_update_web_content():
    """
    Update RAG knowledge base with fresh web content (news, Wikipedia, etc.)
    Expects JSON: {"topics": ["topic1", "topic2", ...]}
    Returns: JSON with success status
    """
    try:
        data = request.json or {}
        topics = data.get('topics', [])
        
        if not topics:
            return jsonify({'error': 'No topics provided'}), 400
        
        if not isinstance(topics, list):
            topics = [topics]
        
        rag = get_rag_service()
        if not rag.enabled:
            return jsonify({'error': 'RAG service not available'}), 503
        
        logger.info(f"🌐 Updating RAG with web content for topics: {topics}")
        success = update_rag_with_custom_content(topics)
        
        if success:
            stats = rag.get_stats()
            return jsonify({
                'status': 'success',
                'topics_processed': len(topics),
                'total_documents': stats['document_count'],
                'message': 'RAG knowledge base updated with web content'
            })
        else:
            return jsonify({
                'error': 'Failed to update RAG with web content'
            }), 500
        
    except Exception as e:
        logger.error(f"Error updating RAG with web content: {e}", exc_info=True)
        return jsonify({'error': 'Failed to update RAG'}), 500


@app.route('/api/rag/update_topic_pack', methods=['POST'])
def rag_update_topic_pack():
    """
    Update RAG using the full configured domain topic pack.
    Expects optional JSON: {"extra_topics": ["topic1", "topic2"]}
    """
    try:
        data = request.json or {}
        extra_topics = data.get('extra_topics', [])

        if extra_topics and not isinstance(extra_topics, list):
            extra_topics = [extra_topics]

        rag = get_rag_service()
        if not rag.enabled:
            return jsonify({'error': 'RAG service not available'}), 503

        logger.info("🚀 Running full RAG topic-pack refresh...")
        result = update_rag_with_topic_pack(extra_topics=extra_topics)
        stats = rag.get_stats()

        return jsonify({
            'status': 'success',
            'topic_pack': result,
            'total_documents': stats['document_count'],
            'message': 'RAG topic-pack refresh completed'
        })

    except Exception as e:
        logger.error(f"Error updating RAG topic pack: {e}", exc_info=True)
        return jsonify({'error': 'Failed to update RAG topic pack'}), 500


@app.route('/api/rag/scheduler/status', methods=['GET'])
def rag_scheduler_status():
    """
    Get RAG update scheduler status
    Returns: JSON with scheduler state, update frequency, last update time
    """
    try:
        status = get_scheduler_status()
        
        logger.info(f"📊 Scheduler status requested")
        
        return jsonify({
            'status': 'success',
            'scheduler': status
        })
        
    except Exception as e:
        logger.error(f"Error retrieving scheduler status: {e}", exc_info=True)
        return jsonify({'error': 'Failed to retrieve scheduler status'}), 500


@app.route('/api/rag/scheduler/force_update', methods=['POST'])
def rag_force_update():
    """
    Force immediate RAG knowledge base update
    Returns: JSON with update result
    """
    try:
        logger.info("🚀 Force update requested")
        success = force_rag_update()
        
        if success:
            stats = get_rag_service().get_stats()
            return jsonify({
                'status': 'success',
                'message': 'RAG knowledge base updated successfully',
                'total_documents': stats['document_count']
            })
        else:
            return jsonify({'error': 'Failed to force update'}), 500
        
    except Exception as e:
        logger.error(f"Error forcing RAG update: {e}", exc_info=True)
        return jsonify({'error': 'Failed to force update'}), 500


@app.route('/api/rag/scheduler/enable', methods=['POST'])
def rag_enable_scheduler():
    """
    Enable RAG auto-updates
    Returns: JSON with success status
    """
    try:
        scheduler = get_rag_update_scheduler()
        
        if not scheduler.enabled:
            logger.info("⏰ Enabling RAG scheduler...")
            # Note: This requires config.py modification for persistence
            scheduler.enabled = True
            scheduler.start()
            return jsonify({'status': 'success', 'message': 'RAG scheduler enabled'})
        else:
            return jsonify({'status': 'success', 'message': 'RAG scheduler already enabled'})
        
    except Exception as e:
        logger.error(f"Error enabling scheduler: {e}", exc_info=True)
        return jsonify({'error': 'Failed to enable scheduler'}), 500


@app.route('/api/rag/scheduler/disable', methods=['POST'])
def rag_disable_scheduler():
    """
    Disable RAG auto-updates
    Returns: JSON with success status
    """
    try:
        scheduler = get_rag_update_scheduler()
        
        if scheduler.enabled:
            logger.info("🛑 Disabling RAG scheduler...")
            scheduler.enabled = False
            scheduler.stop()
            return jsonify({'status': 'success', 'message': 'RAG scheduler disabled'})
        else:
            return jsonify({'status': 'success', 'message': 'RAG scheduler already disabled'})
        
    except Exception as e:
        logger.error(f"Error disabling scheduler: {e}", exc_info=True)
        return jsonify({'error': 'Failed to disable scheduler'}), 500


@app.route('/api/speech', methods=['POST'])
def generate_speech():
    """
    Generate speech audio from text using Google Translate TTS
    Backend processes the TTS - frontend never sees the TTS process
    
    Expects JSON: {"text": "text to speak", "language": "en"}
    Returns: Binary audio data (MP3)
    """
    try:
        data = request.json
        
        # Validate input
        if not data or 'text' not in data:
            return jsonify({'error': 'Text is required'}), 400
        
        text = data.get('text', '').strip()
        language = _normalize_language_code(data.get('language', 'en'))
        
        if not text:
            return jsonify({'error': 'Text cannot be empty'}), 400
        
        logger.info(f"📥 Speech request received - language param: {language}")
        
        # Language mapping for Google Translate TTS
        # All mapped languages have native Google TTS support
        google_lang_map = {
            # International
            'en': 'en',
            
            # Major Indian languages
            'hi': 'hi', 'ta': 'ta', 'te': 'te', 'kn': 'kn', 'ml': 'ml', 'mr': 'mr',
            'gu': 'gu', 'pa': 'pa', 'bn': 'bn',
            
            # Other languages with TTS support
            'ur': 'ur',
        }
        
        google_lang = google_lang_map.get(language, 'en')
        logger.info(f"📥 Mapped {language} → {google_lang}")
        
        # Clean text: remove code blocks and extra formatting
        cleaned_text = text
        cleaned_text = cleaned_text.replace('```', ' ')
        cleaned_text = cleaned_text.replace('`', '')
        cleaned_text = cleaned_text.replace('\n\n', '. ')
        cleaned_text = cleaned_text.replace('\n', ' ')
        cleaned_text = ' '.join(cleaned_text.split())
        
        # Call Google Translate TTS API
        google_tts_url = 'https://translate.google.com/translate_tts'
        params = {
            'ie': 'UTF-8',
            'client': 'tw-ob',
            'q': cleaned_text,
            'tl': google_lang
        }
        
        logger.info(f"🔊 Generating speech: text={cleaned_text[:50]}... | lang={language} | google_lang={google_lang}")
        logger.debug(f"📡 Google TTS URL params: {params}")
        
        try:
            # Make request to Google Translate TTS API
            response = requests.get(google_tts_url, params=params, timeout=10)
            logger.info(f"✅ Google TTS response status: {response.status_code}")
            response.raise_for_status()
            
            # Return audio as binary stream with proper headers
            return Response(
                response.content,
                mimetype='audio/mpeg',
                headers={
                    'Content-Disposition': 'inline',
                    'Cache-Control': 'public, max-age=86400'
                }
            )
            
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ Error calling Google TTS API: {e} | Status: {getattr(e.response, 'status_code', 'N/A')}")
            return jsonify({'error': 'Failed to generate speech', 'details': str(e)}), 500
        
    except Exception as e:
        logger.error(f"Error in speech endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/memory/<user_id>', methods=['GET'])
def get_memory(user_id):
    """Debug endpoint: inspect persisted conversation memory for a user.

    Returns the last N messages (user + assistant) for the given user_id,
    where N is based on CONVERSATION_HISTORY_SIZE.
    """
    try:
        history = memory_db.get_history(user_id, config.CONVERSATION_HISTORY_SIZE)
        return jsonify({
            'user_id': user_id,
            'count': len(history),
            'messages': history,
        })
    except Exception as e:
        logger.error(f"Error reading memory for {user_id}: {e}", exc_info=True)
        return jsonify({'error': 'Failed to read memory'}), 500


@app.route('/api/memory/profile/<user_id>', methods=['GET'])
def get_memory_profile(user_id):
    """Debug endpoint: inspect persisted personal profile memory for a user."""
    try:
        facts = memory_db.get_user_profile_facts(user_id)
        summary = memory_db.get_user_profile_summary(user_id)
        return jsonify({
            'user_id': user_id,
            'fact_count': len(facts),
            'facts': facts,
            'summary': summary,
        })
    except Exception as e:
        logger.error(f"Error reading profile memory for {user_id}: {e}", exc_info=True)
        return jsonify({'error': 'Failed to read profile memory'}), 500


@app.route('/api/chat_stream', methods=['POST'])
def chat_stream():
    """
    Streaming chat endpoint
    Expects JSON: {"message": "user message", "language": "en", "user_id": "optional", "chat_mode": "general"}
    Returns: SSE stream of JSON chunks
    """
    try:
        data = request.json
        if not data or 'message' not in data:
            return jsonify({'error': 'Message is required'}), 400
        
        user_message = data.get('message', '').strip()
        language = _normalize_language_code(data.get('language', 'en'))
        user_id = data.get('user_id', 'default')
        chat_mode = data.get('chat_mode', 'general')
        model_override = data.get('model_override')
        fallback_models = data.get('fallback_models')
        
        if not user_message:
            return jsonify({'error': 'Message cannot be empty'}), 400
        
        logger.info(f"Received streaming request: {user_message[:50]}... (language: {language}, mode: {chat_mode})")

        def stream_orchestrated_chunks():
            """Stream orchestrated response in JSON lines for frontend compatibility."""
            result = orchestrator.handle_query(
                user_message,
                language=language,
                user_id=user_id,
                chat_mode=chat_mode,
                model_override=model_override,
                fallback_models=fallback_models,
            )

            actions = result.get('actions', [])
            sources = result.get('web_search_sources', [])
            if actions:
                yield json.dumps({'actions': actions}) + "\n"
            if sources:
                yield json.dumps({'sources': sources}) + "\n"

            response_text = result.get('response', '')
            if not response_text:
                return

            chunk_size = 200
            for i in range(0, len(response_text), chunk_size):
                chunk = response_text[i:i + chunk_size]
                yield json.dumps({'content': chunk}) + "\n"

        return Response(stream_orchestrated_chunks(), mimetype='text/event-stream')
        
    except Exception as e:
        logger.error(f"Error in chat_stream endpoint: {e}", exc_info=True)
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/process_text', methods=['POST'])

def process_text():
    """Compatibility endpoint for text chat"""
    try:
        data = request.json
        user_message = data.get('text', '').strip()
        language = _normalize_language_code(data.get('language', 'en'))  # Get language from frontend
        user_id = data.get('user_id', 'default')
        
        if not user_message:
            return jsonify({'error': 'Message is required'}), 400
        
        logger.info(f"Received text request: {user_message[:50]}... (language: {language})")
        
        # Get AI response with correct language (now returns tuple)
        ai_response, search_sources = llm.get_response(user_message, language, user_id)
        
        # Generate TTS audio using gTTS
        from gtts import gTTS
        import base64
        import io
        try:
            # Map language codes
            lang_map = {'en': 'en', 'hi': 'hi', 'kn': 'kn', 'te': 'te', 'ta': 'ta',
                        'ml': 'ml', 'mr': 'mr', 'bn': 'bn', 'gu': 'gu', 'pa': 'hi', 'ur': 'ur'}
            tts_lang = lang_map.get(language, 'en')
            
            # Generate TTS
            tts = gTTS(text=ai_response, lang=tts_lang, slow=False, timeout=5)
            audio_fp = io.BytesIO()
            tts.write_to_fp(audio_fp)
            audio_fp.seek(0)
            audio_base64 = base64.b64encode(audio_fp.read()).decode('utf-8')
            audio_mime = 'audio/mpeg'
        except Exception as e:
            logger.error(f"TTS generation failed: {e}")
            audio_base64 = None
            audio_mime = None
        # Return format expected by frontend
        return jsonify({
            'response_text': ai_response,
            'detected_language': language,
            'user_language': language,
            'audio_response': audio_base64,
            'audio_mime': audio_mime,
            'web_search_sources': search_sources
        })
                
    except Exception as e:
        logger.error(f"Error in process_text: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/process_audio', methods=['POST'])
def process_audio():
    """Process audio input - transcribe and chat (multilingual)"""
    try:
        # Check if audio file is present
        if 'audio' not in request.files:
            return jsonify({'error': 'No audio file provided'}), 400
        
        audio_file = request.files['audio']
        
        if audio_file.filename == '':
            return jsonify({'error': 'Empty audio file'}), 400
        
        logger.info(f"Received audio file: {audio_file.filename}")
        
        # Import STT service
        from stt_service import STTService
        stt = STTService()
        
        # Get language hint from request
        raw_language_hint = (request.form.get('language') or '').strip()
        language_hint = _normalize_language_code(raw_language_hint) if raw_language_hint else None
            
        logger.info(f"Processing audio with language hint: {language_hint}")
        
        # Transcribe audio
        transcribed_text, detected_language = stt.transcribe(audio_file, language=language_hint)
        
        if not transcribed_text:
            return jsonify({'error': 'Could not transcribe audio'}), 400
        
        logger.info(f"Transcribed ({detected_language}): {transcribed_text}")
        
        # Get AI response in the detected language (now returns tuple)
        user_id = request.form.get('user_id', 'default')
        ai_response, search_sources = llm.get_response(transcribed_text, detected_language, user_id)
        
        # Return both transcription and response
        # Generate TTS audio for the response
        from gtts import gTTS
        import base64
        import io
        try:
            lang_map = {'en': 'en', 'hi': 'hi', 'kn': 'kn', 'te': 'te', 'ta': 'ta',
                        'ml': 'ml', 'mr': 'mr', 'bn': 'bn', 'gu': 'gu', 'pa': 'hi', 'ur': 'ur'}
            tts_lang = lang_map.get(detected_language, 'en')
            
            tts = gTTS(text=ai_response, lang=tts_lang, slow=False, timeout=5)
            audio_fp = io.BytesIO()
            tts.write_to_fp(audio_fp)
            audio_fp.seek(0)
            audio_base64 = base64.b64encode(audio_fp.read()).decode('utf-8')
            audio_mime = 'audio/mpeg'
        except Exception as e:
            logger.error(f"TTS generation failed: {e}")
            audio_base64 = None
            audio_mime = None
        return jsonify({
            'response_text': ai_response,
            'detected_language': detected_language,
            'user_language': detected_language,
            'audio_response': audio_base64,
            'audio_mime': audio_mime,
            'web_search_sources': search_sources
        })
        
    except Exception as e:
        logger.error(f"Error in process_audio: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/tts_only', methods=['POST'])
def tts_only():
    """Generate TTS audio for any text (multilingual)"""
    try:
        data = request.json
        text = data.get('text', '').strip()
        language = _normalize_language_code(data.get('language', 'en'))
        speed = data.get('speed', 1.0)
        
        if not text:
            return jsonify({'error': 'Text is required'}), 400
        
        logger.info(f"TTS request: {text[:50]}... (lang: {language}, speed: {speed})")
        
        # Use gTTS for multilingual support
        from gtts import gTTS
        import base64
        import io
        
        # Map language codes to gTTS supported codes
        lang_map = {
            'en': 'en',
            'hi': 'hi',
            'kn': 'kn',  # Kannada
            'te': 'te',  # Telugu
            'ta': 'ta',  # Tamil
            'ml': 'ml',  # Malayalam
            'mr': 'mr',  # Marathi
            'bn': 'bn',  # Bengali
            'gu': 'gu',  # Gujarati
            'ur': 'ur',  # Urdu
            'pa': 'pa',  # Punjabi (not supported by gTTS, fallback to Hindi)
        }

        tts_lang = lang_map.get(language, 'en')
        if language == 'pa':  # Punjabi not supported, use Hindi
            tts_lang = 'hi'
            logger.warning(f"Punjabi TTS not supported, using Hindi")

        # Generate TTS
        tts = gTTS(text=text, lang=tts_lang, slow=False, timeout=5)
        
        # Save to BytesIO
        audio_fp = io.BytesIO()
        tts.write_to_fp(audio_fp)
        audio_fp.seek(0)
        
        # Encode to base64
        audio_base64 = base64.b64encode(audio_fp.read()).decode('utf-8')
        
        return jsonify({
            'audio_response': audio_base64,
            'audio_mime': 'audio/mpeg'
        })
        
    except Exception as e:
        logger.error(f"TTS error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ─── AUTHENTICATION & USER MANAGEMENT ───────────────────────────
@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register new user"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        
        if not username or not email or not password:
            return jsonify({'error': 'All fields required'}), 400
        
        user_id, error_or_token = auth_service.register(username, email, password)
        if not user_id:
            return jsonify({'error': error_or_token}), 400
        
        return jsonify({
            'user_id': user_id,
            'token': error_or_token,
            'message': 'Registration successful'
        }), 201
        
    except Exception as e:
        logger.error(f"Registration error: {e}")
        return jsonify({'error': 'Registration failed'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login user"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'error': 'Username and password required'}), 400
        
        user_id, token, error = auth_service.login(username, password)
        if error:
            return jsonify({'error': error}), 401
        
        return jsonify({
            'user_id': user_id,
            'token': token,
            'message': 'Login successful'
        }), 200
        
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({'error': 'Login failed'}), 500

@app.route('/api/auth/verify', methods=['GET'])
@require_auth
def verify_token():
    """Verify if token is valid"""
    try:
        return jsonify({
            'valid': True,
            'user_id': request.user_id
        }), 200
    except Exception as e:
        logger.error(f"Verify error: {e}")
        return jsonify({'error': 'Token verification failed'}), 500

@app.route('/api/profile', methods=['GET'])
@require_auth
def get_profile():
    """Get user profile and statistics"""
    try:
        user_id = request.user_id
        stats = db.get_user_stats(user_id)
        
        return jsonify({
            'user_id': user_id,
            'stats': stats
        }), 200
    except Exception as e:
        logger.error(f"Profile error: {e}")
        return jsonify({'error': 'Failed to get profile'}), 500

@app.route('/api/conversations', methods=['GET'])
@require_auth
def get_conversations():
    """Get all conversations for user"""
    try:
        user_id = request.user_id
        conversations = db.get_conversations(user_id)
        
        # Format for frontend
        formatted = []
        for conv in conversations:
            formatted.append({
                'id': conv['id'],
                'title': conv['title'],
                'language': conv['language'],
                'created_at': conv['created_at'],
                'updated_at': conv['updated_at']
            })
        
        return jsonify({
            'conversations': formatted,
            'count': len(formatted)
        }), 200
    except Exception as e:
        logger.error(f"Error fetching conversations: {e}")
        return jsonify({'error': 'Failed to fetch conversations'}), 500

@app.route('/api/summarize', methods=['POST'])
def summarize():
    """Generate a one-line summary for chat title"""
    try:
        data = request.json
        user_message = data.get('user_message', '').strip()
        ai_response = data.get('ai_response', '').strip()
        
        if not user_message and not ai_response:
            return jsonify({'summary': 'New Chat'}), 200
        
        # Create prompt for summary
        prompt = f"""Generate a SHORT one-line title (max 6 words) that summarizes this conversation. 
Only output the title, nothing else. No quotes, no explanation.

User: {user_message[:200]}
Assistant: {ai_response[:200]}"""
        
        # Use Groq to generate summary
        from groq import Groq
        client = Groq(api_key=config.GROQ_API_KEY)
        
        response = client.messages.create(
            model=config.GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=50
        )
        
        summary = response.choices[0].message.content.strip()
        # Clean up any quotes
        summary = summary.replace('"', '').replace("'", '')
        
        logger.info(f"Chat summary: {summary}")
        return jsonify({'summary': summary}), 200
        
    except Exception as e:
        logger.error(f"Summarize error: {e}", exc_info=True)
        return jsonify({'summary': 'New Chat', 'error': str(e)}), 200

# Register blueprints
app.register_blueprint(chat_management_bp)

def _validate_api_configuration():
    """
    Validate API key configuration at startup.
    Logs warnings if critical APIs are not configured.
    """
    logger.info("🔍 Validating API configuration...")
    logger.info(f"📋 LLM_PROVIDER: {config.LLM_PROVIDER}")
    
    issues = []
    
    # Check if running in OLLAMA-ONLY mode
    if config.LLM_PROVIDER == 'ollama_only':
        logger.info("✅ OLLAMA-ONLY MODE - Groq API not required")
        if config.OLLAMA_ENABLED:
            logger.info(f"✅ Ollama enabled at {config.OLLAMA_API_URL} with model: {config.OLLAMA_MODEL}")
        else:
            issues.append("❌ OLLAMA_ENABLED=False but LLM_PROVIDER=ollama_only")
    elif config.LLM_PROVIDER == 'deepseek_local':
        logger.info("✅ DEEPSEEK-LOCAL MODE - no external API keys required")
        logger.info(f"   Model         : {config.DEEPSEEK_MODEL_NAME}")
        logger.info(f"   Max new tokens: {config.DEEPSEEK_MAX_NEW_TOKENS}")
        logger.info(f"   Temperature   : {config.DEEPSEEK_TEMPERATURE}")
    else:
        # Standard mode - check Groq
        if not config.GROQ_API_KEY:
            issues.append("❌ GROQ_API_KEY not set - LLM responses will fail")
        elif config.GROQ_API_KEY.startswith("your_") or len(config.GROQ_API_KEY) < 20:
            issues.append("⚠️  GROQ_API_KEY looks invalid or placeholder")
        else:
            logger.info("✅ GROQ_API_KEY configured")
        
        # Check Ollama as fallback
        if not config.OLLAMA_ENABLED:
            logger.info("ℹ️  Ollama disabled - will rely on Groq/OpenAI")
    
    # Check OpenAI as fallback
    if not config.OPENAI_API_KEY:
        logger.info("ℹ️  OPENAI_API_KEY not set - OpenAI fallback unavailable")
    
    # Check Serper for search
    if not config.SERPER_API_KEY:
        logger.info("ℹ️  SERPER_API_KEY not set - Web search unavailable")
    
    if issues:
        logger.warning("⚠️  API Configuration Issues Detected:")
        for issue in issues:
            logger.warning(f"  {issue}")
        logger.warning("\n📝 To fix this:")
        logger.warning("  1. Open backend/.env")
        logger.warning("  2. Get a valid API key from https://console.groq.com")
        logger.warning("  3. Update GROQ_API_KEY=your_key_here")
        logger.warning("  4. Restart the server\n")

if __name__ == '__main__':
    _validate_api_configuration()
    logger.info(f"🚀 Starting server on http://localhost:{config.PORT}")
    logger.info("✨ Clean chatbot ready!")
    app.run(host='0.0.0.0', port=config.PORT, debug=config.DEBUG)
